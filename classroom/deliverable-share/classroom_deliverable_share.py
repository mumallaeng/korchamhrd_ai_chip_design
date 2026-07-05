#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
import tempfile
import unicodedata
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any
from xml.etree import ElementTree as ET


XLSX_NS = {
    "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}

FOLDER_MIME = "application/vnd.google-apps.folder"
DEFAULT_BASE_COMMENT = (
    "{url} 편집 권한으로 공유했습니다. 제출물을 확인하고, 편집이 필요한 경우에는 직접 수정해주세요."
)
PLACEHOLDER_URL = "{URL}"
CURRENT_CONFIG: dict[str, Any] = {}
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive"]
CLASSROOM_READONLY_SCOPES = [
    "https://www.googleapis.com/auth/classroom.courses.readonly",
    "https://www.googleapis.com/auth/classroom.rosters.readonly",
    "https://www.googleapis.com/auth/classroom.student-submissions.students.readonly",
    "https://www.googleapis.com/auth/classroom.profile.emails",
]


@dataclass
class ArtifactSpec:
    key: str
    label: str
    match_tokens: list[str]
    expected_kind: str
    allowed_extensions: list[str]
    filename_regex: str
    required_min_count: int


@dataclass
class CandidateCheck:
    name: str
    kind: str
    kind_ok: bool
    extension_ok: bool
    filename_ok: bool


@dataclass
class ArtifactResult:
    key: str
    label: str
    status: str
    candidate_count: int
    valid_count: int
    required_min_count: int
    bad_kind_count: int
    bad_extension_count: int
    bad_filename_count: int
    source_code_zip_count: int
    candidate_names: list[str]
    notes: list[str]


def nfc(value: str) -> str:
    return unicodedata.normalize("NFC", value)


def load_config(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    data.setdefault("drive_sync_root", "")
    data.setdefault("course_id", "")
    data.setdefault("course_work_id", "")
    data.setdefault("pilot_students", [])
    data.setdefault("output_dir", "")
    data.setdefault("top_level_student_name_regex", r"^(?P<student>[^_]+)_")
    data.setdefault("base_comment_template", DEFAULT_BASE_COMMENT)
    data.setdefault("google_oauth_mode", "local-server")
    data.setdefault("google_classroom_token_path", "")
    data.setdefault("excluded_students", [])
    return data


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def read_token_scopes(path: Path) -> set[str]:
    if not path.exists():
        return set()
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return set()
    scopes = payload.get("scopes", [])
    if isinstance(scopes, list):
        return {str(scope) for scope in scopes}
    return set()


def column_letters_to_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    index = 0
    for ch in letters:
        index = index * 26 + (ord(ch.upper()) - ord("A") + 1)
    return index - 1


def read_xlsx_sheet(path: Path, sheet_name: str) -> list[dict[str, str]]:
    shared_strings: list[str] = []
    with zipfile.ZipFile(path) as archive:
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for si in root.findall("a:si", XLSX_NS):
                text = "".join(t.text or "" for t in si.iterfind(".//a:t", XLSX_NS))
                shared_strings.append(text)

        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        sheets = workbook.find("a:sheets", XLSX_NS)
        if sheets is None:
            raise RuntimeError(f"No sheets found in {path}")

        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        relmap = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}

        target = None
        for sheet in sheets:
            if sheet.attrib.get("name") == sheet_name:
                rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                if rel_id:
                    target = "xl/" + relmap[rel_id]
                    break
        if target is None:
            raise RuntimeError(f"Sheet {sheet_name!r} not found in {path}")

        sheet_root = ET.fromstring(archive.read(target))

    rows: list[dict[int, str]] = []
    for row in sheet_root.findall(".//a:sheetData/a:row", XLSX_NS):
        values: dict[int, str] = {}
        for cell in row.findall("a:c", XLSX_NS):
            cell_ref = cell.attrib.get("r", "")
            column_index = column_letters_to_index(cell_ref)
            cell_type = cell.attrib.get("t")
            value_node = cell.find("a:v", XLSX_NS)
            value = ""
            if value_node is not None and value_node.text is not None:
                value = value_node.text
                if cell_type == "s":
                    value = shared_strings[int(value)]
            values[column_index] = value
        if values:
            rows.append(values)

    if not rows:
        return []

    header_row = rows[0]
    max_index = max(header_row)
    headers = [header_row.get(i, "").strip() for i in range(max_index + 1)]

    records: list[dict[str, str]] = []
    for row in rows[1:]:
        record: dict[str, str] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = row.get(index, "").strip()
        if any(value for value in record.values()):
            records.append(record)
    return records


def artifact_specs(config: dict[str, Any]) -> list[ArtifactSpec]:
    specs: list[ArtifactSpec] = []
    for item in config["expected_artifacts"]:
        specs.append(
            ArtifactSpec(
                key=item["key"],
                label=item["label"],
                match_tokens=[nfc(token) for token in item["match_tokens"]],
                expected_kind=item["expected_kind"],
                allowed_extensions=[ext.lower() for ext in item.get("allowed_extensions", [])],
                filename_regex=item["filename_regex"],
                required_min_count=int(item.get("required_min_count", 1)),
            )
        )
    return specs


def roster_lookup(config: dict[str, Any]) -> dict[str, list[dict[str, str]]]:
    roster_config = config["roster"]
    rows = read_xlsx_sheet(Path(roster_config["xlsx_path"]), roster_config["sheet_name"])
    by_name: dict[str, list[dict[str, str]]] = defaultdict(list)
    name_column = roster_config["student_name_column"]
    for row in rows:
        student_name = nfc(row.get(name_column, "").strip())
        if student_name:
            by_name[student_name].append(row)
    return by_name


def parse_top_level_student_name(name: str, pattern: str) -> str | None:
    match = re.match(pattern, nfc(name))
    if not match:
        return None
    return nfc(match.group("student"))


def scan_top_level_entries(config: dict[str, Any]) -> dict[str, dict[str, list[Path]]]:
    deliverable_root = Path(config["deliverable_root"])
    pattern = config["top_level_student_name_regex"]
    grouped: dict[str, dict[str, list[Path]]] = defaultdict(lambda: {"dirs": [], "files": [], "unknown": []})

    for child in sorted(deliverable_root.iterdir()):
        if child.name == ".DS_Store":
            continue
        student_name = parse_top_level_student_name(child.name, pattern)
        if not student_name:
            grouped["__UNPARSEABLE__"]["unknown"].append(child)
            continue
        if child.is_dir():
            grouped[student_name]["dirs"].append(child)
        else:
            grouped[student_name]["files"].append(child)
    return grouped


def classify_kind(path: Path) -> str:
    return "dir" if path.is_dir() else "file"


def build_pattern(spec: ArtifactSpec, student_name: str, slug: str) -> re.Pattern[str]:
    regex = spec.filename_regex.replace("{student}", re.escape(student_name)).replace(
        "{slug}", re.escape(slug)
    )
    return re.compile(regex)


def inspect_artifact(
    spec: ArtifactSpec,
    student_name: str,
    slug: str,
    children: list[Path],
) -> ArtifactResult:
    pattern = build_pattern(spec, student_name, slug)
    candidates: list[CandidateCheck] = []

    for child in children:
        name = nfc(child.name)
        if name.startswith("~$"):
            continue
        if not any(token in name for token in spec.match_tokens):
            continue
        kind = classify_kind(child)
        extension = child.suffix.lower() if child.is_file() else ""
        kind_ok = kind == spec.expected_kind
        extension_ok = True if spec.expected_kind == "dir" else extension in spec.allowed_extensions
        filename_ok = bool(pattern.fullmatch(name))
        candidates.append(
            CandidateCheck(
                name=name,
                kind=kind,
                kind_ok=kind_ok,
                extension_ok=extension_ok,
                filename_ok=filename_ok,
            )
        )

    valid_count = sum(
        1
        for candidate in candidates
        if candidate.kind_ok and candidate.extension_ok and candidate.filename_ok
    )
    bad_kind_count = sum(1 for candidate in candidates if not candidate.kind_ok)
    bad_extension_count = sum(1 for candidate in candidates if not candidate.extension_ok)
    bad_filename_count = sum(1 for candidate in candidates if not candidate.filename_ok)
    source_code_zip_count = sum(
        1
        for candidate in candidates
        if spec.key == "source_code" and candidate.kind == "file" and candidate.name.lower().endswith(".zip")
    )

    notes: list[str] = []
    if not candidates:
        status = "missing"
        notes.append("candidate_missing")
    elif valid_count >= spec.required_min_count:
        status = "pass"
    else:
        status = "invalid"
        if valid_count == 0:
            notes.append("no_valid_candidate")
        else:
            notes.append("insufficient_valid_candidate_count")

    if source_code_zip_count:
        notes.append("source_code_zip_detected")
    if bad_kind_count:
        notes.append("kind_mismatch_detected")
    if bad_extension_count:
        notes.append("extension_mismatch_detected")
    if bad_filename_count:
        notes.append("filename_mismatch_detected")

    return ArtifactResult(
        key=spec.key,
        label=spec.label,
        status=status,
        candidate_count=len(candidates),
        valid_count=valid_count,
        required_min_count=spec.required_min_count,
        bad_kind_count=bad_kind_count,
        bad_extension_count=bad_extension_count,
        bad_filename_count=bad_filename_count,
        source_code_zip_count=source_code_zip_count,
        candidate_names=[candidate.name for candidate in candidates],
        notes=notes,
    )


def build_additional_comments(results: list[ArtifactResult]) -> list[str]:
    comments: list[str] = []
    missing_labels: list[str] = []
    bad_filename_labels: list[str] = []
    bad_format_labels: list[str] = []
    source_code_zip_detected = False

    def render_labels(labels: list[str]) -> str:
        return ", ".join(f"'{label}'" for label in labels)

    for result in results:
        if result.valid_count >= result.required_min_count:
            continue
        if result.source_code_zip_count:
            source_code_zip_detected = True
            continue
        if result.candidate_count == 0:
            missing_labels.append(result.label)
        elif 0 < result.valid_count < result.required_min_count:
            missing_labels.append(result.label)
        if result.bad_filename_count:
            bad_filename_labels.append(result.label)
        if result.bad_kind_count or result.bad_extension_count:
            bad_format_labels.append(result.label)

    if source_code_zip_detected:
        comments.append("산출물 '소스코드'가 압축 파일로 제출되었습니다.")
    if missing_labels:
        comments.append(f"필수 산출물인 {render_labels(missing_labels)}가 누락되었습니다.")
    if bad_filename_labels:
        comments.append(
            f"산출물 {render_labels(bad_filename_labels)}의 파일명이 제출 포맷에 맞지 않습니다."
        )
    if bad_format_labels:
        comments.append(
            f"산출물 {render_labels(bad_format_labels)}의 포맷이 제출 형식에 맞지 않습니다."
        )
    return comments


def build_base_comment(template: str, share_url: str | None) -> str:
    return template.format(url=share_url or PLACEHOLDER_URL)


def split_additional_comments(value: str) -> list[str]:
    return [item.strip() for item in value.split("||") if item.strip()]


def build_submission_followup_comments(
    submission_state: str,
    share_url: str,
) -> list[str]:
    comments: list[str] = []
    if submission_state == "CREATED":
        comments.append("클래스룸 제출 상태가 미제출로 확인됩니다.")
    if not share_url.strip():
        comments.append("공유할 제출물 폴더가 로컬에서 확인되지 않았습니다.")
    return comments


def build_private_comment_draft(
    base_comment: str,
    additional_comments: str,
    submission_state: str = "",
    share_url: str = "",
) -> str:
    lines: list[str] = []
    if share_url.strip() and base_comment.strip():
        lines.append(base_comment.strip())
    lines.extend(build_submission_followup_comments(submission_state, share_url))
    lines.extend(split_additional_comments(additional_comments))
    return "\n".join(lines)


def resolve_output_dir(config: dict[str, Any], override: str | None) -> Path:
    if override:
        return Path(override)
    output_dir = config.get("output_dir")
    if output_dir:
        return Path(output_dir)
    return Path.cwd() / "classroom-deliverable-share-output"


def drive_relative_path(path: Path, drive_sync_root: str) -> str:
    root = Path(drive_sync_root)
    try:
        return str(path.resolve().relative_to(root.resolve()))
    except Exception:
        return ""


def extract_first_date_token(value: str) -> str | None:
    match = re.search(r"(\d{8})", value)
    if match:
        return match.group(1)
    return None


def normalize_basic_name(name: str) -> str:
    return nfc(name).rstrip()


def maybe_redecode_zip_name(value: str) -> str:
    if not value:
        return value
    if any("\uac00" <= ch <= "\ud7a3" for ch in value):
        return value
    if not any(ord(ch) > 127 for ch in value):
        return value
    try:
        raw = value.encode("cp437")
    except UnicodeEncodeError:
        return value
    for encoding in ("cp949", "euc-kr", "utf-8"):
        try:
            decoded = raw.decode(encoding)
        except UnicodeDecodeError:
            continue
        if any("\uac00" <= ch <= "\ud7a3" for ch in decoded):
            return decoded
    return value


def normalize_zip_member_path(member_name: str) -> Path:
    posix_path = PurePosixPath(maybe_redecode_zip_name(member_name))
    cleaned_parts: list[str] = []
    for part in posix_path.parts:
        if part in {"", ".", "/"}:
            continue
        if part == "..":
            raise RuntimeError(f"Unsafe zip member path: {member_name}")
        cleaned_parts.append(normalize_basic_name(part))
    return Path(*cleaned_parts)


def archive_target_dir_for_zip(zip_path: Path) -> Path:
    base_name = normalize_basic_name(zip_path.name)
    if base_name.lower().endswith(".zip"):
        base_name = base_name[:-4]
    base_name = re.sub(r"\s+\(\d+\)$", "", base_name)
    return zip_path.with_name(base_name)


def flatten_single_root_directory(target_dir: Path) -> None:
    children = [child for child in target_dir.iterdir() if child.name != ".DS_Store"]
    if len(children) != 1 or not children[0].is_dir():
        return
    root_dir = children[0]
    for child in list(root_dir.iterdir()):
        child.rename(target_dir / child.name)
    root_dir.rmdir()


def extract_zip_archive_to_dir(zip_path: Path, target_dir: Path) -> dict[str, Any]:
    if target_dir.exists():
        return {
            "source_zip": str(zip_path),
            "target_dir": str(target_dir),
            "status": "skipped_existing_dir",
            "detail": "",
        }

    temp_parent = target_dir.parent
    temp_parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(dir=str(temp_parent), prefix=f".extract-{zip_path.stem}-") as temp_dir_str:
        temp_dir = Path(temp_dir_str)
        with zipfile.ZipFile(zip_path) as archive:
            for info in archive.infolist():
                relative_path = normalize_zip_member_path(info.filename)
                if not relative_path.parts:
                    continue
                destination = temp_dir / relative_path
                if info.is_dir():
                    destination.mkdir(parents=True, exist_ok=True)
                    continue
                destination.parent.mkdir(parents=True, exist_ok=True)
                with archive.open(info) as src, destination.open("wb") as dst:
                    shutil.copyfileobj(src, dst)

        temp_target = temp_dir / target_dir.name
        temp_target.mkdir(parents=True, exist_ok=True)
        for child in list(temp_dir.iterdir()):
            if child == temp_target:
                continue
            shutil.move(str(child), str(temp_target / child.name))
        flatten_single_root_directory(temp_target)
        temp_target.rename(target_dir)

    return {
        "source_zip": str(zip_path),
        "target_dir": str(target_dir),
        "status": "extracted",
        "detail": "",
    }


def build_extract_missing_zips_plan(
    config: dict[str, Any],
    scope: str,
    students_csv: str | None,
) -> list[dict[str, Any]]:
    student_rows, _, _ = assemble_student_rows(
        config=config,
        scope=scope,
        students_csv=students_csv,
    )
    plan_rows: list[dict[str, Any]] = []
    for row in student_rows:
        if row.get("roster_status") != "matched":
            continue
        source_zip = row.get("source_zip", "").strip()
        extracted_dir = row.get("extracted_dir", "").strip()
        if not source_zip or extracted_dir:
            continue
        zip_path = Path(source_zip)
        target_dir = archive_target_dir_for_zip(zip_path)
        plan_rows.append(
            {
                "student_name": row["student_name"],
                "classroom_email": row.get("classroom_email", ""),
                "source_zip": source_zip,
                "target_dir": str(target_dir),
                "status": "planned",
                "detail": "",
            }
        )
    return plan_rows


def canonicalize_direct_child_name(
    child: Path,
    student_name: str,
    assignment_slug: str,
    submission_date: str | None,
) -> str | None:
    normalized_name = normalize_basic_name(child.name)
    if normalized_name.startswith("~$"):
        return None
    artifact_date = extract_first_date_token(normalized_name) or submission_date
    if not artifact_date:
        return None

    suffix = Path(normalized_name).suffix.lower()
    if child.is_dir():
        if "소스코드" in normalized_name:
            return f"{student_name}_{artifact_date}_{assignment_slug}_소스코드"
        return None

    if "일지" in normalized_name and suffix == ".pdf":
        return f"{student_name}_{artifact_date}_{assignment_slug}_일지.pdf"
    if "소스코드" in normalized_name and suffix == ".zip":
        return f"{student_name}_{artifact_date}_{assignment_slug}_소스코드.zip"
    if suffix == ".pptx":
        return f"{student_name}_{artifact_date}_{assignment_slug}_발표자료.pptx"
    if suffix == ".xlsx":
        return f"{student_name}_{artifact_date}_{assignment_slug}_일정표.xlsx"
    if "완료보고서" in normalized_name and suffix in {".pdf", ".docx"}:
        return f"{student_name}_{artifact_date}_{assignment_slug}_완료보고서{suffix}"
    return None


def build_normalization_plan(
    config: dict[str, Any],
    scope: str,
    students_csv: str | None,
) -> list[dict[str, Any]]:
    deliverable_root = Path(config["deliverable_root"])
    pattern = config["top_level_student_name_regex"]
    selected_names = resolve_selected_students(config, scope, students_csv)
    if scope == "all" and not students_csv:
        selected_children = [child for child in sorted(deliverable_root.iterdir()) if child.name != ".DS_Store"]
    elif selected_names:
        selected_children = []
        for child in sorted(deliverable_root.iterdir()):
            if child.name == ".DS_Store":
                continue
            student_name = parse_top_level_student_name(child.name, pattern)
            if student_name and student_name in selected_names:
                selected_children.append(child)
    else:
        selected_children = [child for child in sorted(deliverable_root.iterdir()) if child.name != ".DS_Store"]

    plan_rows: list[dict[str, Any]] = []
    assignment_slug = config["assignment_slug"]

    for top_level_child in selected_children:
        student_name = parse_top_level_student_name(top_level_child.name, pattern)
        submission_date = extract_first_date_token(nfc(top_level_child.name))
        descendants = []
        if top_level_child.is_dir():
            descendants = [path for path in top_level_child.rglob("*") if path.name != ".DS_Store"]
        descendants.sort(key=lambda path: len(path.parts), reverse=True)
        for child in descendants:
            current_name = child.name
            target_name = normalize_basic_name(current_name)
            reasons: list[str] = []

            if target_name != current_name:
                reasons.append("nfc_or_trailing_space_normalized")

            if student_name and child.parent == top_level_child:
                canonical_name = canonicalize_direct_child_name(
                    child=child,
                    student_name=student_name,
                    assignment_slug=assignment_slug,
                    submission_date=submission_date,
                )
                if canonical_name and canonical_name != target_name:
                    target_name = canonical_name
                    reasons.append("artifact_name_canonicalized")

            if target_name == current_name:
                continue

            plan_rows.append(
                {
                    "student_name": student_name,
                    "current_path": str(child),
                    "target_path": str(child.with_name(target_name)),
                    "reasons": " ; ".join(reasons),
                }
            )

        top_level_target_name = normalize_basic_name(top_level_child.name)
        if top_level_target_name != top_level_child.name:
            plan_rows.append(
                {
                    "student_name": student_name or "",
                    "current_path": str(top_level_child),
                    "target_path": str(top_level_child.with_name(top_level_target_name)),
                    "reasons": "nfc_or_trailing_space_normalized",
                }
            )
    return plan_rows


def apply_normalization_plan(plan_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for row in plan_rows:
        current_path = Path(row["current_path"])
        target_path = Path(row["target_path"])
        status = "skipped"
        detail = ""

        if not current_path.exists():
            status = "missing_source"
        elif target_path.exists() and target_path != current_path:
            try:
                if current_path.samefile(target_path):
                    status = "samefile_equivalent"
                else:
                    status = "conflict_existing_target"
            except OSError:
                status = "conflict_existing_target"
        else:
            try:
                current_path.rename(target_path)
                status = "renamed"
            except OSError as exc:
                status = "rename_failed"
                detail = str(exc)

        result = dict(row)
        result["status"] = status
        result["detail"] = detail
        results.append(result)
    return results


def maybe_auto_fix_names(
    config: dict[str, Any],
    output_dir: Path,
    scope: str,
    students_csv: str | None,
    enabled: bool,
) -> None:
    if not enabled:
        return
    plan_rows = build_normalization_plan(
        config=config,
        scope=scope,
        students_csv=students_csv,
    )
    write_csv(output_dir / "normalization-plan.csv", plan_rows)
    results = apply_normalization_plan(plan_rows)
    write_csv(output_dir / "normalization-results.csv", results)


def resolve_selected_students(config: dict[str, Any], scope: str, students_csv: str | None) -> set[str]:
    if students_csv:
        return {nfc(item.strip()) for item in students_csv.split(",") if item.strip()}
    if scope == "pilot":
        return {nfc(student) for student in config.get("pilot_students", [])}
    return set()


def excluded_students(config: dict[str, Any]) -> set[str]:
    return {nfc(student) for student in config.get("excluded_students", []) if student}


def roster_email(row: dict[str, str], config: dict[str, Any]) -> str:
    column = config["roster"]["classroom_email_column"]
    return row.get(column, "").strip()


def assemble_student_rows(
    config: dict[str, Any],
    scope: str,
    students_csv: str | None,
    share_url_map: dict[str, str] | None = None,
    share_status_map: dict[str, str] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    by_name = roster_lookup(config)
    grouped = scan_top_level_entries(config)
    selected_names = resolve_selected_students(config, scope, students_csv)

    if scope == "all" and not students_csv:
        all_students = {name for name in grouped.keys() if name != "__UNPARSEABLE__"}
    elif selected_names:
        all_students = selected_names
    else:
        all_students = {name for name in grouped.keys() if name != "__UNPARSEABLE__"}
    all_students -= excluded_students(config)

    all_students |= {name for name in by_name.keys() if name in selected_names}
    all_students -= excluded_students(config)

    rows: list[dict[str, Any]] = []
    validation_rows: list[dict[str, Any]] = []

    specs = artifact_specs(config)
    slug = config["assignment_slug"]
    share_url_map = share_url_map or {}
    share_status_map = share_status_map or {}

    summary = {
        "selected_students": 0,
        "roster_missing": 0,
        "email_missing": 0,
        "share_candidates": 0,
        "validation_warnings": 0,
        "validation_failures": 0,
    }

    for student_name in sorted(all_students):
        top_level = grouped.get(student_name, {"dirs": [], "files": [], "unknown": []})
        roster_rows = by_name.get(student_name, [])
        roster_status = "matched"
        notes: list[str] = []
        classroom_email = ""
        if not roster_rows:
            roster_status = "missing"
            summary["roster_missing"] += 1
            notes.append("roster_missing")
        elif len(roster_rows) > 1:
            roster_status = "duplicate"
            notes.append("roster_duplicate")
        else:
            classroom_email = roster_email(roster_rows[0], config)
            if not classroom_email:
                notes.append("classroom_email_missing")
                summary["email_missing"] += 1

        extracted_dir = top_level["dirs"][0] if len(top_level["dirs"]) == 1 else None
        if len(top_level["dirs"]) > 1:
            notes.append("multiple_extracted_dirs")
        elif extracted_dir is None:
            notes.append("extracted_dir_missing")

        source_zip = top_level["files"][0] if len(top_level["files"]) == 1 else None
        if len(top_level["files"]) > 1:
            notes.append("multiple_source_archives")
        elif source_zip is None:
            notes.append("source_archive_missing")

        selected = scope == "all" or student_name in selected_names or not selected_names
        if not selected:
            continue
        summary["selected_students"] += 1

        artifact_results: list[ArtifactResult] = []
        validation_status = "blocked"
        share_candidate = bool(extracted_dir and classroom_email and roster_status == "matched")
        if share_candidate:
            summary["share_candidates"] += 1

        if extracted_dir and extracted_dir.is_dir():
            children = [child for child in sorted(extracted_dir.iterdir()) if child.name != ".DS_Store"]
            artifact_results = [inspect_artifact(spec, student_name, slug, children) for spec in specs]
            if any(result.status == "invalid" for result in artifact_results):
                validation_status = "fail"
                summary["validation_failures"] += 1
            elif any(result.status == "missing" for result in artifact_results):
                validation_status = "fail"
                summary["validation_failures"] += 1
            elif any(result.status == "warn" for result in artifact_results):
                validation_status = "warn"
                summary["validation_warnings"] += 1
            else:
                validation_status = "pass"
        else:
            notes.append("validation_skipped_no_extracted_dir")

        additional_comments = build_additional_comments(artifact_results)
        share_url = share_url_map.get(student_name, "")
        share_status = share_status_map.get(student_name, "pending")
        base_comment = build_base_comment(config["base_comment_template"], share_url or None)

        row = {
            "student_name": student_name,
            "selected_scope": scope,
            "roster_status": roster_status,
            "classroom_email": classroom_email,
            "extracted_dir": str(extracted_dir) if extracted_dir else "",
            "source_zip": str(source_zip) if source_zip else "",
            "drive_relative_path": drive_relative_path(extracted_dir, config["drive_sync_root"]) if extracted_dir else "",
            "share_candidate": "yes" if share_candidate else "no",
            "share_status": share_status,
            "share_url": share_url,
            "validation_status": validation_status,
            "base_comment": base_comment,
            "additional_comments": " || ".join(additional_comments),
            "notes": " ; ".join(notes),
        }
        rows.append(row)

        for result in artifact_results:
            validation_rows.append(
                {
                    "student_name": student_name,
                    "artifact_key": result.key,
                    "artifact_label": result.label,
                    "status": result.status,
                    "candidate_count": result.candidate_count,
                    "valid_count": result.valid_count,
                    "required_min_count": result.required_min_count,
                    "bad_kind_count": result.bad_kind_count,
                    "bad_extension_count": result.bad_extension_count,
                    "bad_filename_count": result.bad_filename_count,
                    "source_code_zip_count": result.source_code_zip_count,
                    "candidate_names": " || ".join(result.candidate_names),
                    "notes": " ; ".join(result.notes),
                }
            )

    return rows, validation_rows, summary


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_private_comment_workbook(path: Path, plan_rows: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to write the private-comment workbook."
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "private_comments"

    headers = [
        "비공개 댓글 복붙용",
        "학생명",
        "클래스룸 이메일",
        "과제명",
        "제출 상태",
        "매칭 상태",
        "검증 상태",
        "공유 상태",
        "제출물 링크",
        "공유 폴더 URL",
        "추가 댓글",
        "API 작성 상태",
    ]
    sheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment

    for row in plan_rows:
        private_comment_draft = row.get("private_comment_draft", "").strip() or "없음"
        sheet.append(
            [
                private_comment_draft,
                row.get("student_name", ""),
                row.get("classroom_email", ""),
                row.get("course_work_title", ""),
                row.get("submission_state", ""),
                row.get("match_status", ""),
                row.get("validation_status", ""),
                row.get("share_status", ""),
                row.get("submission_alternate_link", ""),
                row.get("share_url", ""),
                row.get("additional_comments", ""),
                row.get("comment_apply_status", ""),
            ]
        )

    widths = {
        "A": 72,
        "B": 14,
        "C": 28,
        "D": 42,
        "E": 14,
        "F": 18,
        "G": 14,
        "H": 18,
        "I": 48,
        "J": 48,
        "K": 40,
        "L": 24,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment

    sheet.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def actionable_private_comment_rows(plan_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        row
        for row in plan_rows
        if row.get("submission_id", "").strip() and row.get("share_url", "").strip()
    ]


def submission_private_comment_rows(plan_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row for row in plan_rows if row.get("submission_id", "").strip()]


def write_summary(path: Path, config: dict[str, Any], summary: dict[str, Any], student_rows: list[dict[str, Any]]) -> None:
    lines = [
        f"# {config['assignment_title']} run summary",
        "",
        f"- selected students: {summary['selected_students']}",
        f"- roster missing: {summary['roster_missing']}",
        f"- email missing: {summary['email_missing']}",
        f"- share candidates: {summary['share_candidates']}",
        f"- validation warnings: {summary['validation_warnings']}",
        f"- validation failures: {summary['validation_failures']}",
        "",
        "## Students",
        "",
    ]
    for row in student_rows:
        lines.append(
            f"- {row['student_name']}: share={row['share_status']}, validation={row['validation_status']}, "
            f"email={row['classroom_email'] or 'MISSING'}, notes={row['notes'] or '-'}"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def emit_reports(
    config: dict[str, Any],
    output_dir: Path,
    scope: str,
    students_csv: str | None,
    share_url_map: dict[str, str] | None = None,
    share_status_map: dict[str, str] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    student_rows, validation_rows, summary = assemble_student_rows(
        config=config,
        scope=scope,
        students_csv=students_csv,
        share_url_map=share_url_map,
        share_status_map=share_status_map,
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "student-roster-resolved.csv", student_rows)
    write_csv(output_dir / "share-plan.csv", student_rows)
    write_csv(output_dir / "comment-queue.csv", student_rows)
    write_csv(output_dir / "validation-report.csv", validation_rows)
    write_summary(output_dir / "run-summary.md", config, summary, student_rows)
    return student_rows, validation_rows, summary


def require_google_modules() -> tuple[Any, Any, Any, Any]:
    try:
        from google.auth.transport.requests import Request
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise RuntimeError(
            "Google API client libraries are not installed. Install the requirements file first."
        ) from exc
    return Request, Credentials, InstalledAppFlow, build


def scopes_satisfied(creds: Any, scopes: list[str]) -> bool:
    existing_scopes = set(getattr(creds, "scopes", []) or [])
    return set(scopes).issubset(existing_scopes)


def google_service(
    credentials_path: Path,
    token_path: Path,
    scopes: list[str],
    service_name: str,
    version: str,
) -> Any:
    Request, Credentials, InstalledAppFlow, build = require_google_modules()
    creds = None
    requested_scopes = set(scopes)
    token_scopes = read_token_scopes(token_path)
    if token_path.exists() and requested_scopes.issubset(token_scopes):
        creds = Credentials.from_authorized_user_file(str(token_path), scopes)
        if creds and not scopes_satisfied(creds, scopes):
            creds = None
    if creds and creds.valid and scopes_satisfied(creds, scopes):
        return build(service_name, version, credentials=creds, cache_discovery=False)
    if creds and creds.expired and creds.refresh_token and scopes_satisfied(creds, scopes):
        creds.refresh(Request())
    if not creds or not creds.valid:
        flow = InstalledAppFlow.from_client_secrets_file(str(credentials_path), scopes)
        oauth_mode = "console"
        if credentials_path.name:
            pass
        if CURRENT_CONFIG.get("google_oauth_mode"):
            oauth_mode = CURRENT_CONFIG["google_oauth_mode"]
        if oauth_mode == "local-server":
            creds = flow.run_local_server(
                port=0,
                open_browser=False,
                authorization_prompt_message=(
                    "Open this URL in your browser and authorize the app:\n{url}"
                ),
            )
        else:
            auth_url, _ = flow.authorization_url(
                access_type="offline",
                include_granted_scopes="true",
                prompt="consent",
            )
            print("Open this URL in your browser and authorize the app:")
            print(auth_url)
            code = input("Paste the authorization code here: ").strip()
            flow.fetch_token(code=code)
            creds = flow.credentials
    token_path.parent.mkdir(parents=True, exist_ok=True)
    token_path.write_text(creds.to_json(), encoding="utf-8")
    return build(service_name, version, credentials=creds, cache_discovery=False)


def drive_service(credentials_path: Path, token_path: Path) -> Any:
    return google_service(credentials_path, token_path, DRIVE_SCOPES, "drive", "v3")


def classroom_service(credentials_path: Path, token_path: Path) -> Any:
    return google_service(credentials_path, token_path, CLASSROOM_READONLY_SCOPES, "classroom", "v1")


def classroom_token_path(config: dict[str, Any], override: str | None = None) -> Path:
    if override:
        return Path(override)
    configured = config.get("google_classroom_token_path", "") or config.get("google_token_path", "")
    if not configured:
        raise RuntimeError(
            "Missing Classroom token path. Set google_classroom_token_path or google_token_path in config."
        )
    return Path(configured)


def drive_query_escape(name: str) -> str:
    return name.replace("\\", "\\\\").replace("'", "\\'")


def drive_name_variants(name: str) -> list[str]:
    variants = [name, nfc(name), unicodedata.normalize("NFD", name)]
    seen: set[str] = set()
    unique: list[str] = []
    for variant in variants:
        if variant not in seen:
            seen.add(variant)
            unique.append(variant)
    return unique


def drive_find_child(service: Any, parent_id: str, child_name: str, want_folder: bool) -> dict[str, Any]:
    queries: list[str] = []
    for variant in drive_name_variants(child_name):
        base = (
            f"trashed = false and '{parent_id}' in parents and name = '{drive_query_escape(variant)}'"
        )
        if want_folder:
            base += f" and mimeType = '{FOLDER_MIME}'"
        queries.append(base)

    matches: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for query in queries:
        response = service.files().list(
            q=query,
            fields="files(id,name,mimeType,webViewLink,parents)",
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            pageSize=50,
        ).execute()
        for item in response.get("files", []):
            if item["id"] not in seen_ids:
                seen_ids.add(item["id"])
                matches.append(item)

    if not matches:
        raise RuntimeError(f"Drive path segment not found under parent {parent_id}: {child_name}")
    if len(matches) > 1:
        raise RuntimeError(f"Drive path segment is ambiguous under parent {parent_id}: {child_name}")
    return matches[0]


def resolve_drive_item_by_local_path(service: Any, local_path: Path, drive_sync_root: Path) -> dict[str, Any]:
    relative_parts = local_path.resolve().relative_to(drive_sync_root.resolve()).parts
    parent_id = "root"
    current: dict[str, Any] | None = None
    for index, part in enumerate(relative_parts):
        want_folder = index < len(relative_parts) - 1 or local_path.is_dir()
        current = drive_find_child(service, parent_id, part, want_folder)
        parent_id = current["id"]
    if current is None:
        raise RuntimeError(f"Could not resolve Drive item for {local_path}")
    return current


def existing_permission(service: Any, file_id: str, classroom_email: str) -> dict[str, Any] | None:
    response = service.permissions().list(
        fileId=file_id,
        fields="permissions(id,emailAddress,role,type)",
        supportsAllDrives=True,
    ).execute()
    for permission in response.get("permissions", []):
        if permission.get("emailAddress", "").lower() == classroom_email.lower():
            return permission
    return None


def apply_drive_permission(
    service: Any,
    file_id: str,
    classroom_email: str,
    target_role: str,
) -> str:
    if target_role not in {"writer", "reader"}:
        raise RuntimeError(f"Unsupported Drive role: {target_role}")

    permission = existing_permission(service, file_id, classroom_email)
    if permission:
        role = permission.get("role", "")
        if target_role == "writer":
            if role in {"writer", "owner", "organizer", "fileOrganizer"}:
                return "already_shared"
            service.permissions().update(
                fileId=file_id,
                permissionId=permission["id"],
                body={"role": "writer"},
                fields="id,role",
                supportsAllDrives=True,
            ).execute()
            return "upgraded_to_writer"

        if role in {"reader", "commenter"}:
            return "already_viewer_shared"
        if role == "owner":
            return "owner_unchanged"
        service.permissions().update(
            fileId=file_id,
            permissionId=permission["id"],
            body={"role": "reader"},
            fields="id,role",
            supportsAllDrives=True,
        ).execute()
        return "downgraded_to_reader"

    service.permissions().create(
        fileId=file_id,
        body={"type": "user", "role": target_role, "emailAddress": classroom_email},
        fields="id",
        sendNotificationEmail=(target_role == "writer"),
        supportsAllDrives=True,
    ).execute()
    if target_role == "writer":
        return "shared_writer_invite_sent"
    return "shared_viewer_without_notification"


def fetch_web_view_link(service: Any, file_id: str) -> str:
    metadata = service.files().get(
        fileId=file_id,
        fields="id,webViewLink",
        supportsAllDrives=True,
    ).execute()
    return metadata.get("webViewLink", "")


def load_existing_share_maps(output_dir: Path) -> tuple[dict[str, str], dict[str, str]]:
    share_plan_rows = read_csv(output_dir / "share-plan.csv")
    share_url_map: dict[str, str] = {}
    share_status_map: dict[str, str] = {}
    for row in share_plan_rows:
        student_name = row.get("student_name", "")
        if not student_name:
            continue
        if row.get("share_url"):
            share_url_map[student_name] = row["share_url"]
        if row.get("share_status"):
            share_status_map[student_name] = row["share_status"]
    return share_url_map, share_status_map


def classroom_list_all(fetch_page: Any, item_key: str) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    page_token: str | None = None
    while True:
        response = fetch_page(page_token)
        items.extend(response.get(item_key, []))
        page_token = response.get("nextPageToken")
        if not page_token:
            break
    return items


def list_classroom_courses(service: Any) -> list[dict[str, Any]]:
    return classroom_list_all(
        lambda page_token: service.courses()
        .list(teacherId="me", pageSize=100, courseStates=["ACTIVE"], pageToken=page_token)
        .execute(),
        "courses",
    )


def list_classroom_coursework(service: Any, course_id: str) -> list[dict[str, Any]]:
    return classroom_list_all(
        lambda page_token: service.courses()
        .courseWork()
        .list(courseId=course_id, pageSize=100, pageToken=page_token)
        .execute(),
        "courseWork",
    )


def list_classroom_students(service: Any, course_id: str) -> list[dict[str, Any]]:
    return classroom_list_all(
        lambda page_token: service.courses()
        .students()
        .list(courseId=course_id, pageSize=100, pageToken=page_token)
        .execute(),
        "students",
    )


def list_classroom_student_submissions(service: Any, course_id: str, course_work_id: str) -> list[dict[str, Any]]:
    return classroom_list_all(
        lambda page_token: service.courses()
        .courseWork()
        .studentSubmissions()
        .list(courseId=course_id, courseWorkId=course_work_id, pageSize=100, pageToken=page_token)
        .execute(),
        "studentSubmissions",
    )


def get_user_profile(service: Any, user_id: str) -> dict[str, Any]:
    return service.userProfiles().get(userId=user_id).execute()


def normalize_title(value: str) -> str:
    return nfc(value).strip()


def resolve_coursework_context(
    service: Any,
    config: dict[str, Any],
    course_id_override: str | None = None,
    course_work_id_override: str | None = None,
) -> dict[str, str]:
    configured_course_id = course_id_override or config.get("course_id", "")
    configured_course_work_id = course_work_id_override or config.get("course_work_id", "")
    assignment_title = normalize_title(config.get("assignment_title", ""))

    if configured_course_work_id and not configured_course_id:
        raise RuntimeError("course_work_id cannot be used without course_id.")

    if configured_course_id and configured_course_work_id:
        course = service.courses().get(id=configured_course_id).execute()
        course_work = (
            service.courses()
            .courseWork()
            .get(courseId=configured_course_id, id=configured_course_work_id)
            .execute()
        )
        return {
            "course_id": configured_course_id,
            "course_name": course.get("name", ""),
            "course_work_id": configured_course_work_id,
            "course_work_title": course_work.get("title", ""),
        }

    if configured_course_id:
        course = service.courses().get(id=configured_course_id).execute()
        course_work_items = list_classroom_coursework(service, configured_course_id)
        matches = [
            work for work in course_work_items if normalize_title(work.get("title", "")) == assignment_title
        ]
        if not matches:
            raise RuntimeError(
                f"No coursework title matched {assignment_title!r} in course {configured_course_id}."
            )
        if len(matches) > 1:
            raise RuntimeError(
                f"Multiple coursework items matched {assignment_title!r} in course {configured_course_id}."
            )
        course_work = matches[0]
        return {
            "course_id": configured_course_id,
            "course_name": course.get("name", ""),
            "course_work_id": course_work["id"],
            "course_work_title": course_work.get("title", ""),
        }

    courses = list_classroom_courses(service)
    matches: list[dict[str, str]] = []
    for course in courses:
        course_id = course["id"]
        for course_work in list_classroom_coursework(service, course_id):
            if normalize_title(course_work.get("title", "")) == assignment_title:
                matches.append(
                    {
                        "course_id": course_id,
                        "course_name": course.get("name", ""),
                        "course_work_id": course_work["id"],
                        "course_work_title": course_work.get("title", ""),
                    }
                )
    if not matches:
        raise RuntimeError(f"No coursework title matched {assignment_title!r} across active courses.")
    if len(matches) > 1:
        rendered = ", ".join(
            f"{match['course_name']}:{match['course_work_title']}({match['course_work_id']})"
            for match in matches
        )
        raise RuntimeError(f"Multiple coursework matches found for {assignment_title!r}: {rendered}")
    return matches[0]


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def build_classroom_comment_plan_rows(
    service: Any,
    config: dict[str, Any],
    student_rows: list[dict[str, Any]],
    context: dict[str, str],
) -> list[dict[str, Any]]:
    submissions = list_classroom_student_submissions(
        service, context["course_id"], context["course_work_id"]
    )
    try:
        students = list_classroom_students(service, context["course_id"])
    except Exception as exc:
        raise RuntimeError(
            "Unable to list Classroom students for this course. Re-authorize with classroom.rosters.readonly."
        ) from exc
    student_profile_by_user_id: dict[str, dict[str, Any]] = {}
    for student in students:
        user_id = student.get("userId", "")
        if user_id:
            student_profile_by_user_id[user_id] = student.get("profile", {})
    local_by_name = {nfc(row["student_name"]): row for row in student_rows}
    local_by_email = {
        row["classroom_email"].strip().lower(): row
        for row in student_rows
        if row.get("classroom_email", "").strip()
    }
    matched_local_names: set[str] = set()
    plan_rows: list[dict[str, Any]] = []

    for submission in submissions:
        user_id = submission.get("userId", "")
        profile = student_profile_by_user_id.get(user_id, {})
        submission_email = profile.get("emailAddress", "").strip()
        submission_name = nfc(profile.get("name", {}).get("fullName", ""))

        local_row = None
        if submission_email:
            local_row = local_by_email.get(submission_email.lower())
        if local_row is None and submission_name:
            local_row = local_by_name.get(submission_name)

        if local_row is not None:
            matched_local_names.add(nfc(local_row["student_name"]))

        base_comment = local_row["base_comment"] if local_row else build_base_comment(DEFAULT_BASE_COMMENT, None)
        additional_comments = local_row["additional_comments"] if local_row else ""
        validation_status = local_row["validation_status"] if local_row else "unknown"
        share_status = local_row["share_status"] if local_row else "unknown"
        share_url = local_row["share_url"] if local_row else ""
        private_comment_draft = build_private_comment_draft(
            base_comment,
            additional_comments,
            submission_state=submission.get("state", ""),
            share_url=share_url,
        )

        plan_rows.append(
            {
                "course_id": context["course_id"],
                "course_name": context["course_name"],
                "course_work_id": context["course_work_id"],
                "course_work_title": context["course_work_title"],
                "student_name": local_row["student_name"] if local_row else submission_name,
                "classroom_email": local_row["classroom_email"] if local_row else submission_email,
                "submission_user_id": user_id,
                "submission_id": submission.get("id", ""),
                "submission_state": submission.get("state", ""),
                "submission_alternate_link": submission.get("alternateLink", ""),
                "submission_name_from_classroom": submission_name,
                "submission_email_from_classroom": submission_email,
                "match_status": "matched" if local_row else "classroom_only_unmatched",
                "share_status": share_status,
                "share_url": share_url,
                "validation_status": validation_status,
                "base_comment": base_comment,
                "additional_comments": additional_comments,
                "private_comment_draft": private_comment_draft,
                "private_comment_api_status": "unavailable_in_official_rest_api_as_of_2026-05-20",
                "comment_apply_status": "blocked_api_unavailable",
            }
        )

    for local_row in student_rows:
        student_name = nfc(local_row["student_name"])
        if student_name in matched_local_names:
            continue
        plan_rows.append(
            {
                "course_id": context["course_id"],
                "course_name": context["course_name"],
                "course_work_id": context["course_work_id"],
                "course_work_title": context["course_work_title"],
                "student_name": local_row["student_name"],
                "classroom_email": local_row["classroom_email"],
                "submission_user_id": "",
                "submission_id": "",
                "submission_state": "",
                "submission_alternate_link": "",
                "submission_name_from_classroom": "",
                "submission_email_from_classroom": "",
                "match_status": "local_only_no_submission_match",
                "share_status": local_row["share_status"],
                "share_url": local_row["share_url"],
                "validation_status": local_row["validation_status"],
                "base_comment": local_row["base_comment"],
                "additional_comments": local_row["additional_comments"],
                "private_comment_draft": build_private_comment_draft(
                    "",
                    local_row["additional_comments"],
                    submission_state="",
                    share_url=local_row["share_url"],
                ),
                "private_comment_api_status": "unavailable_in_official_rest_api_as_of_2026-05-20",
                "comment_apply_status": "blocked_missing_submission_match",
            }
        )
    return plan_rows


def command_scan(args: argparse.Namespace) -> int:
    global CURRENT_CONFIG
    config = load_config(Path(args.config))
    CURRENT_CONFIG = config
    output_dir = resolve_output_dir(config, args.output_dir)
    auto_fix = args.auto_fix_names or config.get("auto_fix_names_before_scan", False)
    maybe_auto_fix_names(
        config=config,
        output_dir=output_dir,
        scope=args.scope,
        students_csv=args.students,
        enabled=auto_fix,
    )
    _, _, summary = emit_reports(
        config=config,
        output_dir=output_dir,
        scope=args.scope,
        students_csv=args.students,
    )
    print(f"scan complete: {output_dir}")
    print("auto_fix_names:", "yes" if auto_fix else "no")
    print(
        "selected={selected_students} share_candidates={share_candidates} "
        "validation_failures={validation_failures} validation_warnings={validation_warnings}".format(**summary)
    )
    return 0


def command_normalize_names(args: argparse.Namespace) -> int:
    global CURRENT_CONFIG
    config = load_config(Path(args.config))
    CURRENT_CONFIG = config
    output_dir = resolve_output_dir(config, args.output_dir)
    plan_rows = build_normalization_plan(
        config=config,
        scope=args.scope,
        students_csv=args.students,
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "normalization-plan.csv", plan_rows)

    if args.apply:
        results = apply_normalization_plan(plan_rows)
        write_csv(output_dir / "normalization-results.csv", results)
        renamed = sum(1 for row in results if row["status"] == "renamed")
        print(f"normalize-names apply complete: {output_dir}")
        print(f"planned={len(plan_rows)} renamed={renamed}")
    else:
        print(f"normalize-names dry-run complete: {output_dir}")
        print(f"planned={len(plan_rows)}")
    return 0


def command_extract_missing_zips(args: argparse.Namespace) -> int:
    global CURRENT_CONFIG
    config = load_config(Path(args.config))
    CURRENT_CONFIG = config
    output_dir = resolve_output_dir(config, args.output_dir)
    plan_rows = build_extract_missing_zips_plan(
        config=config,
        scope=args.scope,
        students_csv=args.students,
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "extract-missing-zips-plan.csv", plan_rows)

    if args.apply:
        results: list[dict[str, Any]] = []
        for row in plan_rows:
            source_zip = Path(row["source_zip"])
            target_dir = Path(row["target_dir"])
            try:
                result = extract_zip_archive_to_dir(source_zip, target_dir)
            except Exception as exc:
                result = {
                    "source_zip": str(source_zip),
                    "target_dir": str(target_dir),
                    "status": "extract_failed",
                    "detail": str(exc),
                }
            result["student_name"] = row["student_name"]
            result["classroom_email"] = row.get("classroom_email", "")
            results.append(result)
        write_csv(output_dir / "extract-missing-zips-results.csv", results)
        extracted = sum(1 for row in results if row["status"] == "extracted")
        print(f"extract-missing-zips apply complete: {output_dir}")
        print(f"planned={len(plan_rows)} extracted={extracted}")
    else:
        print(f"extract-missing-zips dry-run complete: {output_dir}")
        print(f"planned={len(plan_rows)}")
    return 0


def command_share_drive(args: argparse.Namespace) -> int:
    global CURRENT_CONFIG
    config = load_config(Path(args.config))
    CURRENT_CONFIG = config
    output_dir = resolve_output_dir(config, args.output_dir)
    auto_fix = args.auto_fix_names or config.get("auto_fix_names_before_scan", False)
    maybe_auto_fix_names(
        config=config,
        output_dir=output_dir,
        scope=args.scope,
        students_csv=args.students,
        enabled=auto_fix,
    )
    student_rows, _, _ = assemble_student_rows(
        config=config,
        scope=args.scope,
        students_csv=args.students,
    )

    credentials_path = Path(args.credentials or config.get("google_client_secret_path", ""))
    token_path = Path(args.token or config.get("google_token_path", ""))
    drive_sync_root = Path(config["drive_sync_root"])
    if not credentials_path:
        raise RuntimeError("Missing Google OAuth client secret path. Set google_client_secret_path or --credentials.")
    if not token_path:
        raise RuntimeError("Missing token path. Set google_token_path or --token.")
    if not drive_sync_root:
        raise RuntimeError("Missing drive_sync_root in config.")

    service = drive_service(credentials_path, token_path)
    share_url_map: dict[str, str] = {}
    share_status_map: dict[str, str] = {}
    target_role = "reader" if args.role == "viewer" else args.role

    for row in student_rows:
        student_name = row["student_name"]
        if row["share_candidate"] != "yes":
            share_status_map[student_name] = "blocked"
            continue

        extracted_dir = Path(row["extracted_dir"])
        classroom_email = row["classroom_email"]
        try:
            drive_item = resolve_drive_item_by_local_path(service, extracted_dir, drive_sync_root)
            share_url = drive_item.get("webViewLink", "") or fetch_web_view_link(service, drive_item["id"])
            share_url_map[student_name] = share_url
            if args.apply:
                status = apply_drive_permission(service, drive_item["id"], classroom_email, target_role)
            else:
                status = "dry_run_resolved"
            share_status_map[student_name] = status
        except Exception as exc:
            share_status_map[student_name] = f"error:{exc}"

    emit_reports(
        config=config,
        output_dir=output_dir,
        scope=args.scope,
        students_csv=args.students,
        share_url_map=share_url_map,
        share_status_map=share_status_map,
    )
    print(f"share-drive complete: {output_dir}")
    print("auto_fix_names:", "yes" if auto_fix else "no")
    print("apply mode:", "yes" if args.apply else "no")
    print("target role:", args.role)
    return 0


def command_classroom_list_courses(args: argparse.Namespace) -> int:
    global CURRENT_CONFIG
    config = load_config(Path(args.config))
    CURRENT_CONFIG = config
    credentials_path = Path(args.credentials or config.get("google_client_secret_path", ""))
    token_path = classroom_token_path(config, args.token)
    service = classroom_service(credentials_path, token_path)
    courses = list_classroom_courses(service)
    output_dir = resolve_output_dir(config, args.output_dir)
    rows = [
        {
            "course_id": course.get("id", ""),
            "name": course.get("name", ""),
            "section": course.get("section", ""),
            "room": course.get("room", ""),
            "enrollment_code": course.get("enrollmentCode", ""),
        }
        for course in courses
    ]
    write_csv(output_dir / "classroom-courses.csv", rows)
    print(f"classroom-list-courses complete: {output_dir}")
    print(f"courses={len(rows)}")
    return 0


def command_classroom_list_coursework(args: argparse.Namespace) -> int:
    global CURRENT_CONFIG
    config = load_config(Path(args.config))
    CURRENT_CONFIG = config
    credentials_path = Path(args.credentials or config.get("google_client_secret_path", ""))
    token_path = classroom_token_path(config, args.token)
    service = classroom_service(credentials_path, token_path)
    course_id = args.course_id or config.get("course_id", "")
    if not course_id:
        raise RuntimeError("Missing course_id. Pass --course-id or set course_id in config.")
    output_dir = resolve_output_dir(config, args.output_dir)
    course_work_items = list_classroom_coursework(service, course_id)
    rows = [
        {
            "course_id": course_id,
            "course_work_id": item.get("id", ""),
            "title": item.get("title", ""),
            "work_type": item.get("workType", ""),
            "state": item.get("state", ""),
            "alternate_link": item.get("alternateLink", ""),
        }
        for item in course_work_items
    ]
    write_csv(output_dir / "classroom-coursework.csv", rows)
    print(f"classroom-list-coursework complete: {output_dir}")
    print(f"coursework={len(rows)}")
    return 0


def command_classroom_build_comment_plan(args: argparse.Namespace) -> int:
    global CURRENT_CONFIG
    config = load_config(Path(args.config))
    CURRENT_CONFIG = config
    output_dir = resolve_output_dir(config, args.output_dir)
    auto_fix = args.auto_fix_names or config.get("auto_fix_names_before_scan", False)
    maybe_auto_fix_names(
        config=config,
        output_dir=output_dir,
        scope=args.scope,
        students_csv=args.students,
        enabled=auto_fix,
    )

    share_url_map, share_status_map = load_existing_share_maps(output_dir)
    if not share_url_map and not share_status_map:
        _, _, _ = emit_reports(
            config=config,
            output_dir=output_dir,
            scope=args.scope,
            students_csv=args.students,
        )
        share_url_map, share_status_map = load_existing_share_maps(output_dir)

    student_rows, _, _ = assemble_student_rows(
        config=config,
        scope=args.scope,
        students_csv=args.students,
        share_url_map=share_url_map,
        share_status_map=share_status_map,
    )

    credentials_path = Path(args.credentials or config.get("google_client_secret_path", ""))
    token_path = classroom_token_path(config, args.token)
    service = classroom_service(credentials_path, token_path)
    context = resolve_coursework_context(
        service=service,
        config=config,
        course_id_override=args.course_id,
        course_work_id_override=args.course_work_id,
    )
    plan_rows = build_classroom_comment_plan_rows(service, config, student_rows, context)
    selected_names = resolve_selected_students(config, args.scope, args.students)
    if args.scope != "all" or selected_names:
        selected_names = selected_names or {nfc(row["student_name"]) for row in student_rows}
        plan_rows = [row for row in plan_rows if nfc(row.get("student_name", "")) in selected_names]
    excluded_names = excluded_students(config)
    if excluded_names:
        plan_rows = [row for row in plan_rows if nfc(row.get("student_name", "")) not in excluded_names]
    write_csv(output_dir / "classroom-comment-plan.csv", plan_rows)
    write_private_comment_workbook(output_dir / "classroom-private-comment-copybook.xlsx", plan_rows)
    write_private_comment_workbook(
        output_dir / "classroom-private-comment-copybook-submitters.xlsx",
        submission_private_comment_rows(plan_rows),
    )
    write_private_comment_workbook(
        output_dir / "classroom-private-comment-copybook-ready.xlsx",
        actionable_private_comment_rows(plan_rows),
    )
    write_json(output_dir / "classroom-context.json", context)
    print(f"classroom-build-comment-plan complete: {output_dir}")
    print(
        f"rows={len(plan_rows)} course_id={context['course_id']} course_work_id={context['course_work_id']}"
    )
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Prepare and optionally share extracted Google Classroom deliverable folders."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    scan_parser = subparsers.add_parser("scan", help="Build local validation and comment queue reports.")
    scan_parser.add_argument("--config", required=True, help="Assignment config JSON path.")
    scan_parser.add_argument("--scope", choices=["pilot", "all"], default="pilot")
    scan_parser.add_argument("--students", help="Comma-separated explicit student names.")
    scan_parser.add_argument("--output-dir", help="Override output directory.")
    scan_parser.add_argument(
        "--auto-fix-names",
        action="store_true",
        help="Apply Windows-safe Hangul normalization and artifact canonicalization before scanning.",
    )
    scan_parser.set_defaults(func=command_scan)

    normalize_parser = subparsers.add_parser(
        "normalize-names",
        help="Build or apply Windows-safe Hangul rename plans inside extracted student folders.",
    )
    normalize_parser.add_argument("--config", required=True, help="Assignment config JSON path.")
    normalize_parser.add_argument("--scope", choices=["pilot", "all"], default="pilot")
    normalize_parser.add_argument("--students", help="Comma-separated explicit student names.")
    normalize_parser.add_argument("--output-dir", help="Override output directory.")
    normalize_parser.add_argument(
        "--apply",
        action="store_true",
        help="Actually rename files and folders. Without this flag, only a rename plan is generated.",
    )
    normalize_parser.set_defaults(func=command_normalize_names)

    extract_parser = subparsers.add_parser(
        "extract-missing-zips",
        help="Extract zip-only student submissions into shareable folders.",
    )
    extract_parser.add_argument("--config", required=True, help="Assignment config JSON path.")
    extract_parser.add_argument("--scope", choices=["pilot", "all"], default="pilot")
    extract_parser.add_argument("--students", help="Comma-separated explicit student names.")
    extract_parser.add_argument("--output-dir", help="Override output directory.")
    extract_parser.add_argument(
        "--apply",
        action="store_true",
        help="Actually extract the missing zip archives. Without this flag, only an extraction plan is generated.",
    )
    extract_parser.set_defaults(func=command_extract_missing_zips)

    share_parser = subparsers.add_parser(
        "share-drive",
        help="Resolve Drive folders for selected students and optionally update Drive sharing permissions.",
    )
    share_parser.add_argument("--config", required=True, help="Assignment config JSON path.")
    share_parser.add_argument("--scope", choices=["pilot", "all"], default="pilot")
    share_parser.add_argument("--students", help="Comma-separated explicit student names.")
    share_parser.add_argument("--output-dir", help="Override output directory.")
    share_parser.add_argument(
        "--auto-fix-names",
        action="store_true",
        help="Apply Windows-safe Hangul normalization and artifact canonicalization before sharing.",
    )
    share_parser.add_argument("--credentials", help="OAuth client secret JSON path.")
    share_parser.add_argument("--token", help="OAuth token JSON path.")
    share_parser.add_argument(
        "--role",
        choices=["writer", "viewer"],
        default="writer",
        help="Drive permission role to enforce for each student's shared folder.",
    )
    share_parser.add_argument(
        "--apply",
        action="store_true",
        help="Actually create or update Drive permissions. Without this flag, only resolve paths.",
    )
    share_parser.set_defaults(func=command_share_drive)

    courses_parser = subparsers.add_parser(
        "classroom-list-courses",
        help="List active Classroom courses for the authenticated teacher account.",
    )
    courses_parser.add_argument("--config", required=True, help="Assignment config JSON path.")
    courses_parser.add_argument("--output-dir", help="Override output directory.")
    courses_parser.add_argument("--credentials", help="OAuth client secret JSON path.")
    courses_parser.add_argument("--token", help="Classroom OAuth token JSON path.")
    courses_parser.set_defaults(func=command_classroom_list_courses)

    coursework_parser = subparsers.add_parser(
        "classroom-list-coursework",
        help="List Classroom coursework items for one course.",
    )
    coursework_parser.add_argument("--config", required=True, help="Assignment config JSON path.")
    coursework_parser.add_argument("--course-id", help="Explicit Classroom course ID.")
    coursework_parser.add_argument("--output-dir", help="Override output directory.")
    coursework_parser.add_argument("--credentials", help="OAuth client secret JSON path.")
    coursework_parser.add_argument("--token", help="Classroom OAuth token JSON path.")
    coursework_parser.set_defaults(func=command_classroom_list_coursework)

    comment_plan_parser = subparsers.add_parser(
        "classroom-build-comment-plan",
        help="Resolve Classroom submissions and produce a private-comment plan CSV.",
    )
    comment_plan_parser.add_argument("--config", required=True, help="Assignment config JSON path.")
    comment_plan_parser.add_argument("--scope", choices=["pilot", "all"], default="pilot")
    comment_plan_parser.add_argument("--students", help="Comma-separated explicit student names.")
    comment_plan_parser.add_argument("--course-id", help="Explicit Classroom course ID override.")
    comment_plan_parser.add_argument("--course-work-id", help="Explicit Classroom course work ID override.")
    comment_plan_parser.add_argument("--output-dir", help="Override output directory.")
    comment_plan_parser.add_argument(
        "--auto-fix-names",
        action="store_true",
        help="Apply Windows-safe Hangul normalization and artifact canonicalization before planning.",
    )
    comment_plan_parser.add_argument("--credentials", help="OAuth client secret JSON path.")
    comment_plan_parser.add_argument("--token", help="Classroom OAuth token JSON path.")
    comment_plan_parser.set_defaults(func=command_classroom_build_comment_plan)

    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return args.func(args)
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
